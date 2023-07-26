import asyncio
import itertools

from deepdiff import DeepDiff
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class DynamicClassMeta(type):
    def __new__(cls, name, bases, attrs):
        # 获取前端传递的字段信息
        fields = attrs.pop('fields', {})

        # 动态添加属性和方法
        for field, value in fields.items():
            attrs[field] = value

        # 创建类
        return super().__new__(cls, name, bases, attrs)


class DynamicClassBase(metaclass=DynamicClassMeta):
    pass


class DiffExcel():
    def __init__(self, target, reference, kwargs):
        self.df1 = pd.read_excel(reference)
        self.df2 = pd.read_excel(target)

        self.dic1 = self.df1.to_dict()
        self.dic2 = self.df2.to_dict()

        self.bom = self.dic1[kwargs['Field 1']]
        self.spare = self.dic2[kwargs['Field 1']]
        self.reference_quntity = self.dic1[kwargs['Field 2']]
        self.target_quntity = self.dic2[kwargs['Field 2']]

        self.new_refereance_quntity = {self.bom[key]: self.reference_quntity[key] for key in self.bom.keys() if key in self.reference_quntity}
        self.new_target_quntity = {self.spare[key]: self.target_quntity[key] for key in self.spare.keys() if key in self.target_quntity}

        self.add = []
        self.remove = []
        self.changed = []
        self.new = []
        self.old = []
        self.coordinates = []

        self.wb = Workbook()
        # 获取默认的工作表
        self.ws = self.wb.active


    async def missing(self, key):
        missing_dict = {}
        if self.bom[key] not in self.spare.values():
            print({"item": key + 1, "PART NUMBER": self.bom[key]})

    def _re(self, string):
        result = re.sub(r"root\['(.*)'\]", r"\1", string)
        return result

    def fill(self, row):
        # 打印列的索引
        for r, row in enumerate(self.df1.itertuples(index=False), start=1):
            # 遍历每一列
            for c, value in enumerate(row, start=1):
                # 设置单元格的值
                self.ws.cell(row=r, column=c, value=value)
                # 设置指定行的背景颜色为黄色
                if r == row + 1:
                    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    self.ws.cell(row=r, column=c).fill = fill

    def diff(self, type=0):
        res = DeepDiff(self.new_refereance_quntity, self.new_target_quntity)
        return res


    def sort(self):
        diff_result = self.diff()

        if 'dictionary_item_added' in diff_result.keys():
            for diff_add in diff_result['dictionary_item_added']:
                data = self._re(diff_add)
                self.add.append(data)
            print(f"The data added in comparison between the current target file and the reference file is;{self.add}")
        else:
            print("no add data")

        if "dictionary_item_removed" in diff_result.keys():
            for diff_remove in diff_result['dictionary_item_removed']:
                if isinstance(diff_remove, str):
                    self.remove.append(self._re(diff_remove))

            print(
                f"The data remove in comparison between the current target file and the reference file is;{self.remove}")
        else:
            print("no remove data")

        if "values_changed" in diff_result.keys():
            for diff_change in diff_result['values_changed']:
                if isinstance(diff_change, str):
                    self.changed.append(self._re(diff_change))
                    self.new.append(diff_result['values_changed'][diff_change]["new_value"])
                    self.old.append(diff_result['values_changed'][diff_change]["old_value"])

            print(
                f"The data changed in comparison between the current target file and the reference file is;{self.changed}")
        else:
            print("no change data")

        return self.add, self.remove, self.changed, self.new, self.old

    async def different_quantity(self, key):
        if self.spare[key] in self.bom.values():
            pass
        else:
            print("reference file has't this PART NUMBER: ", {"item": key + 1, "PART NUMBER": self.bom[key]})

    def organize_quntity(self, key):
        new_dict = {self.bom[key]: self.reference_quntity[key] for key in self.bom.keys() if key in self.reference_quntity}
        return new_dict

    async def main(self):
        tasks = [asyncio.ensure_future(self.missing(key)) for key in  self.bom.keys()]
        await asyncio.wait(tasks)


    def run(self):
        # loop = asyncio.get_event_loop()
        # loop.run_until_complete(self.main())
        res = DeepDiff(self.new_refereance_quntity, self.new_target_quntity)
        res2 = DeepDiff(self.new_refereance_unit, self.new_target_unit)

if __name__ == '__main__':
    # DiffExcel().test()
    import random
    # dict1 = {'a': 1, 'b': 2, 'c': 3}
    # dict2 = {'b': 1, 'a': 4, 'c': 3}
    #
    # res = DeepDiff(dict1, dict2)
    # print(res)
    start = 0  # 起始值
    stop = 10  # 终止值
    step = 2  # 步长

    result = [range(start, stop, step)]
    print([range(-1000,1000,12)])